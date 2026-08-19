#!/usr/bin/env python3
"""Generate WFM Watch security review response Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"c:\Users\SM5587\OneDrive - Zebra Technologies\Daily_Work\WFMControlM\docs\WFM-Watch-Security-Review-Response.xlsx"

ROWS = [
    {
        "no": 1,
        "security_question": "Have you considered using a dedicated secrets management service (e.g., HashiCorp Vault) instead of storing the master CONFIG_ENCRYPTION_KEY in a server environment variable?",
        "current_state_reply": "Client DB2 passwords are encrypted at rest with AES-256-GCM using CONFIG_ENCRYPTION_KEY from the server environment. Optional Keeper Secrets Manager SDK exists for DB2 password override but is not required. Most other AppConfig secrets (JWT, SMTP, SSH) are stored in SQLite and are not encrypted on write today.",
        "plan_of_action": "Phase 3: Encrypt all AppConfig secrets at rest using existing CONFIG_ENCRYPTION_KEY. Document env-var injection via deployment platform (Docker/K8s secrets). Vault/Keeper remains optional future enhancement if org mandates external secrets store — not required for app operation.",
        "implementable": "Partial (app)",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 3",
        "status": "Planned",
    },
    {
        "no": 2,
        "security_question": "What is the automated procedure to safely rotate the master CONFIG_ENCRYPTION_KEY and re-encrypt all client secrets without causing an application outage?",
        "current_state_reply": "No automated CONFIG_ENCRYPTION_KEY rotation exists today. Client DB2 passwords can be bulk-updated via API; a one-time encrypt migration script exists. Rotating the master key without re-encryption would break decryption.",
        "plan_of_action": "Phase 3: Create documented runbook + script: (1) decrypt with old key, (2) re-encrypt with new key for all Client.db2Password and encrypted AppConfig rows, (3) rolling restart. Schedule maintenance window for first rotation.",
        "implementable": "Yes (app/scripts)",
        "owner": "Development + Ops",
        "priority": "Medium",
        "phase": "Phase 3",
        "status": "Planned",
    },
    {
        "no": 3,
        "security_question": "How does the application handle decrypted client DB2 passwords in memory? Are they immediately scrubbed or zeroed out after a connection is established?",
        "current_state_reply": "Passwords are decrypted in Node, passed to child JVM via DB2_PASS_OVERRIDE env var for JDBC connect, then connection is closed. API never returns passwords. No explicit memory scrubbing — strings remain in heap until GC.",
        "plan_of_action": "Phase 3: Minimize exposure — clear env vars after jjs exit, avoid logging credentials, shorten password lifetime in memory. Note: full zero-memory scrubbing is not realistically achievable in Node.js (immutable strings).",
        "implementable": "Partial (limited in Node.js)",
        "owner": "Development",
        "priority": "Low",
        "phase": "Phase 3",
        "status": "Planned",
    },
    {
        "no": 4,
        "security_question": "Will the application server be deployed in an isolated private subnet (DMZ/VPC) to separate it from the public-facing reverse proxy?",
        "current_state_reply": "Not enforced by application code. Deployment architecture is an infrastructure decision. Docker/nginx config assumes reverse proxy can terminate traffic separately from API.",
        "plan_of_action": "Infra/Ops: Deploy app server in private subnet; expose only reverse proxy (nginx) to operators. Document in production deployment guide. Out of application codebase scope.",
        "implementable": "No — Infra only",
        "owner": "Infrastructure / Network",
        "priority": "High",
        "phase": "Production deployment",
        "status": "Ops decision required",
    },
    {
        "no": 5,
        "security_question": "Is there strict egress firewall filtering on the App Server to ensure it can ONLY initiate connections to registered client IP addresses on approved ports (e.g., 22, 50001)?",
        "current_state_reply": "Application initiates outbound SSH (22), DB2 (50000), and SMTP (25/587) based on configured clients but does not enforce egress rules.",
        "plan_of_action": "Infra/Ops: Configure host/VPC egress firewall allowlist from registered client records (IPs + ports 22, 50000, 25/587). App maintains client IP registry; firewall policy is ops responsibility.",
        "implementable": "No — Infra only",
        "owner": "Infrastructure / Network",
        "priority": "High",
        "phase": "Production deployment",
        "status": "Ops decision required",
    },
    {
        "no": 6,
        "security_question": "Are the DB2 connections (port 50000) and SMTP connections (port 25) strictly enforcing TLS/SSL to prevent credentials and job data traversing the network in plaintext?",
        "current_state_reply": "DB2: Assessed Aug 2026 — client DB2 servers on port 50030 do not support TLS; JDBC sslConnection=true times out; plain JDBC succeeds (IBM security mechanism 3). App cannot enforce DB2 transport TLS without DBA enabling server-side SSL. DB2 JDBC SSL app flags removed. Mitigation: private network + egress firewall allowlist to registered client IPs. SMTP: Opt-in TLS via secrets.smtpTlsEnabled (requireTLS on 587, cert validation configurable). Legacy dev mode disables TLS for Mailpit only. App HTTPS: infra.trustProxy + infra.requireHttps enforce HTTPS for operator UI/API when behind nginx/LB.",
        "plan_of_action": "Production go-live: (1) Set secrets.smtpTlsEnabled=true, smtpPort=587, production relay. (2) Set infra.trustProxy=true, infra.requireHttps=true; TLS cert on nginx/LB. (3) Infra: egress allowlist ports 22, 50030, 587. (4) Future: DBA enables DB2 TLS + truststore; then re-add JDBC SSL support. See docs/WFM-Watch-Security-Review-TLS-Response.md.",
        "implementable": "Partial — SMTP & app HTTPS (app); DB2 TLS requires DBA/Infra",
        "owner": "Development + Ops + DBA",
        "priority": "High",
        "phase": "Production deployment / DBA dependency",
        "status": "Partially mitigated — DB2 pending server-side TLS",
    },
    {
        "no": 7,
        "security_question": "Will Multi-Factor Authentication (MFA) or Enterprise SSO be enforced for operators logging into the dashboard?",
        "current_state_reply": "Login is username/password only (bcrypt + JWT). SSH TOTP exists for app-server connections, not dashboard login. Open registration endpoint POST /api/auth/register exists without auth.",
        "plan_of_action": "Phase 1: Disable or restrict open registration. Phase 4 (if required): Add in-app TOTP MFA for operators (no external API — reuse SSH TOTP pattern). Enterprise SSO would require external IdP integration — out of scope unless mandated.",
        "implementable": "Partial (in-app TOTP possible; SSO needs IdP)",
        "owner": "Development + Identity team",
        "priority": "High",
        "phase": "Phase 1 / Phase 4",
        "status": "Planned",
    },
    {
        "no": 8,
        "security_question": "Does the login endpoint implement rate-limiting, temporary account lockouts, or CAPTCHA to prevent brute-force and credential-stuffing attacks?",
        "current_state_reply": "Implemented (Aug 2026). POST /api/auth/login rate-limited to 10 failed attempts per IP per 15 minutes (express-rate-limit). Successful logins not counted. Returns HTTP 429 when exceeded. Login attempts logged with IP.",
        "plan_of_action": "Implemented. See docs/WFM-Watch-Security-Review-Login-Rate-Limit-Response.md.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 1",
        "status": "Implemented",
    },
    {
        "no": 9,
        "security_question": "Does the API enforce data-level authorization to prevent IDOR, ensuring a user for Client A cannot access data for Client B?",
        "current_state_reply": "JWT required on /api/*. Write ops use requirePermission RBAC. Read APIs for payroll, unprocessed-punch, and db-monitor lack backend function checks — frontend-only RBAC can be bypassed. Ack endpoints accept spoofable req.body.userId.",
        "plan_of_action": "Phase 1: Add requirePermission on GET payroll, unprocessed-punch, db-monitor routes. Use JWT user identity for all ack/suppress actions instead of request body userId.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 1",
        "status": "Planned",
    },
    {
        "no": 10,
        "security_question": "Are all state-changing API endpoints, especially in the Admin UI, protected with anti-CSRF tokens?",
        "current_state_reply": "API uses Bearer JWT in Authorization header, not cookies. No CSRF tokens. Classic CSRF risk is low for Bearer-header APIs; XSS + localStorage token theft is the greater concern.",
        "plan_of_action": "Phase 2: If JWT moved to HttpOnly cookies (#15), add SameSite=Strict cookies + optional CSRF token for state-changing requests. No action needed while using Bearer header from memory/cookie without cross-site cookie auth.",
        "implementable": "Yes (if cookies adopted)",
        "owner": "Development",
        "priority": "Low",
        "phase": "Phase 2 (conditional)",
        "status": "Planned",
    },
    {
        "no": 11,
        "security_question": "How does the application validate the log path from client cron files to prevent Path Traversal attacks (e.g., reading /etc/shadow)?",
        "current_state_reply": "Implemented (Phase 1 go-live): Log paths from cron are validated via validateRemoteLogPath() in backend/src/utils/remote-path.ts — allowlist under infra.sshWfmPathPrefix (default /mount/RWS4/) and /mount/backup, reject .. and shell metacharacters, require .log/.txt/.out extension. Invalid paths rejected at ingest (sync) and use (log check, GET /api/jobs/:id/log-tail). SSH stat/tail use shellQuote(). Cron entry file path validated before cat.",
        "plan_of_action": "Done for go-live. Monitor for clients using non-standard log roots; extend allowlist via infra.sshWfmPathPrefix if needed.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 1 (go-live)",
        "status": "Implemented",
    },
    {
        "no": 12,
        "security_question": "Is all data ingested from client systems strictly sanitized and HTML-encoded before being rendered on the dashboard or in emails to prevent Stored XSS?",
        "current_state_reply": "React escapes UI text by default. Notify email templates use escHtml(). Generic alert emails and some metadata fields are not HTML-escaped. No centralized server-side sanitization library.",
        "plan_of_action": "Phase 2: Apply escHtml() to all dynamic email content in alert-service. Audit UI for dangerouslySetInnerHTML (none today). Consider DOMPurify if rich HTML ever added.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 2",
        "status": "Planned",
    },
    {
        "no": 13,
        "security_question": "What mechanism parses remote cron files? Does it prevent any part of the remote file from being evaluated as executable code (Command Injection)?",
        "current_state_reply": "Cron read via SSH cat (path validated), parsed locally with regex — not executed locally. pgrep/grep keys sanitized (sanitizePgrepSearchTerm, sanitizeGrepKey) and shellQuote() applied in sync-service. Remaining Phase 2: audit other SSH command builders.",
        "plan_of_action": "Phase 1 (partial): Sanitized pgrep/grep/tail/stat in sync-service and log-tail API. Phase 2: Full audit of all SSH exec paths.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 1 (partial) / Phase 2",
        "status": "In progress",
    },
    {
        "no": 14,
        "security_question": "How does the application actively revoke a JWT session token if a user's account is compromised or they log out, before natural expiry?",
        "current_state_reply": "Implemented (Aug 2026). JWTs include jti + tokenVersion (tv). RevokedToken denylist on logout. authMiddleware + WebSocket check revocation. Admin can revoke all user/master sessions. Password change/deactivation bumps tokenVersion.",
        "plan_of_action": "Implemented. See docs/WFM-Watch-Security-Review-JWT-Revocation-Response.md. HttpOnly cookies remain separate (#15).",
        "implementable": "Implemented",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 2",
        "status": "Implemented",
    },
    {
        "no": 15,
        "security_question": "How is the session token stored in the user's browser? Is it in a secure, HttpOnly cookie to prevent theft via XSS?",
        "current_state_reply": "Implemented (Aug 2026). JWT in HttpOnly SameSite=Strict cookie (wfm_session). Secure flag when HTTPS/production. Frontend uses withCredentials; no localStorage token. Session restored via GET /api/auth/me. WebSocket uses cookie on handshake.",
        "plan_of_action": "Implemented. See docs/WFM-Watch-Security-Review-HttpOnly-Cookie-Response.md.",
        "implementable": "Implemented",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 2",
        "status": "Implemented",
    },
    {
        "no": 16,
        "security_question": "What algorithm is used to sign JWTs? Is it a strong algorithm (e.g., RS256), and is the app protected from signature-stripping (none algorithm) attacks?",
        "current_state_reply": "Implemented (Aug 2026). HS256 with explicit algorithms: ['HS256'] on verify and algorithm: 'HS256' on sign. Blocks none/algorithm-switching attacks. RS256 not used (symmetric secret).",
        "plan_of_action": "Implemented. See docs/WFM-Watch-Security-Review-JWT-Algorithm-Response.md. JWT secret encryption remains #1.",
        "implementable": "Implemented",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 1",
        "status": "Implemented",
    },
    {
        "no": 17,
        "security_question": "Are the SSH accounts used by WFM Watch restricted with a forced command or restricted shell so they can ONLY read specific files?",
        "current_state_reply": "Mitigated (Aug 2026). App: read-only command allowlist (validateRemoteCommand) before every SSH exec, path allowlists (#11). Password auth (+ optional TOTP). ForceCommand/restricted shell is client sshd config — see docs/SSH-Account-Hardening-Client-Servers.md.",
        "plan_of_action": "App-side command guardrail implemented (ssh-client.ts). Client ops deploy sshd Match User + ForceCommand at each app-server go-live.",
        "implementable": "Partial (app + client sshd config)",
        "owner": "Development + Client Ops",
        "priority": "High",
        "phase": "Phase 2 + Client deployment",
        "status": "Mitigated (shared)",
    },
    {
        "no": 18,
        "security_question": "Are the DB2 user accounts restricted at the database level to have ONLY SELECT permissions on required monitoring tables?",
        "current_state_reply": "All app SQL is hardcoded SELECT. Custom query endpoint blocks non-SELECT keywords (naive filter). App assumes DB2 user is read-only — not verified programmatically.",
        "plan_of_action": "DBA/Client: Grant SELECT-only on required tables per client. Dev: Harden custom query guard (stricter parsing, disable in prod if not needed). Document minimum DB2 privileges.",
        "implementable": "Partial (app guard + DBA grants)",
        "owner": "DBA + Development",
        "priority": "High",
        "phase": "Phase 2 + Client deployment",
        "status": "Planned",
    },
    {
        "no": 19,
        "security_question": "Is there a hardcoded minimum limit for DB polling interval and a limit on concurrent connections to prevent DoS on client databases?",
        "current_state_reply": "Implemented (Aug 2026). Config save rejects DB polling intervals below 5 minutes and db2QueryConcurrency/db2PoolMax outside 1–10. Runtime getInt clamps out-of-range values. Keys: polling.batchRefreshMins, punchRefreshMins, backgroundPollingMins, dbMonitorSyncMins, batchCacheTtlMins, punchCacheTtlMins, engine.db2QueryConcurrency, infra.db2PoolMax.",
        "plan_of_action": "Implemented. See backend/src/utils/config-limits.ts.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 3",
        "status": "Implemented",
    },
    {
        "no": 20,
        "security_question": "Are strict connection and read timeouts enforced for all outbound SSH and DB2 queries to prevent hung connections stalling the monitoring cycle?",
        "current_state_reply": "SSH connect timeout 15s, remote commands 10–60s, jjs child 120s, pool acquire 30s. JDBC layer has no explicit connect/query timeout. Some jobs.ts SSH exec paths lack command timeout.",
        "plan_of_action": "Phase 3: Add JDBC connect/query timeouts in DB2Connector.js. Add command timeout to untimed SSH exec in jobs.ts. Document all timeout config keys.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 3",
        "status": "Planned",
    },
    {
        "no": 21,
        "security_question": "Is there an alert aggregation or deduplication mechanism to prevent alert storms (e.g., 5,000 emails) causing SMTP DoS?",
        "current_state_reply": "Partially implemented: one open escalated alert per client with count updates; per-rule cooldownMinutes; email notifyCooldownMins (default 60). No cross-rule global deduplication.",
        "plan_of_action": "Phase 2 (optional tighten): Review cooldown defaults for prod. Add cross-rule fingerprint dedup if needed. Current mechanism largely adequate — document behavior for security reviewers.",
        "implementable": "Mostly done — minor enhancements possible",
        "owner": "Development",
        "priority": "Low",
        "phase": "Phase 2 (optional)",
        "status": "Mostly implemented",
    },
    {
        "no": 22,
        "security_question": "Are logs forwarded in real-time to a centralized, immutable logging server (SIEM) to prevent tampering of local logs?",
        "current_state_reply": "Winston logs to local rotating files (error.log, combined.log) and console. No SIEM integration or structured JSON format for ingestion.",
        "plan_of_action": "Infra/Ops: Deploy log shipper (Filebeat/Fluent Bit) to forward logs to SIEM. Dev (optional): Add structured JSON log format in Winston for easier ingestion — no external API required in app.",
        "implementable": "No — Infra primary; optional app JSON logs",
        "owner": "Infrastructure + Development",
        "priority": "Medium",
        "phase": "Production deployment",
        "status": "Ops decision required",
    },
    {
        "no": 23,
        "security_question": "When an alert is acknowledged, is the exact operator identity and timestamp securely recorded in an immutable audit trail?",
        "current_state_reply": "acknowledgedBy and acknowledgedAt stored on AlertEvent, EscalatedAlert, UnprocPunchAlert. AuditLog table used for job CRUD only. acknowledgedBy comes from request body (spoofable), not JWT.",
        "plan_of_action": "Phase 1: Record req.user.id from JWT + timestamp + IP in AuditLog for all ack/suppress actions. Remove reliance on client-supplied userId.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 1",
        "status": "Planned",
    },
    {
        "no": 24,
        "security_question": "Are the audit logs scrubbed to ensure no sensitive data (passwords, session tokens, API keys) is accidentally written to log files?",
        "current_state_reply": "Implemented (Aug 2026). Global Winston redactFormat on all transports — JWT, Bearer, JDBC creds, env secrets, sensitive JSON keys. Client db2Password still redacted in update logs. Remote log-tail API scrubs lines before response. Admin UI masks secrets.",
        "plan_of_action": "Implemented. See docs/WFM-Watch-Security-Review-Log-Scrubbing-Response.md.",
        "implementable": "Implemented",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 1",
        "status": "Implemented",
    },
    {
        "no": 25,
        "security_question": "Do the Payroll and Unprocessed Punch screens retrieve or display sensitive PII? If so, is it masked or redacted before sending to the browser?",
        "current_state_reply": "Payroll: SELECT * FROM TA_UNIT_PAY_STATUS with few columns excluded — unit-level operational data (unit IDs, statuses, dates), not employee names/SSN by design. Unprocessed Punch: aggregates only (count, timestamps) — no row-level PII. Backend lacks PAYROLL_VIEW permission check.",
        "plan_of_action": "Phase 2: Replace SELECT * with explicit column allowlist for payroll. Enforce PAYROLL_VIEW / UNPROC_PUNCH_VIEW on backend GET routes (#9). Document columns returned.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 2",
        "status": "Planned",
    },
    {
        "no": 26,
        "security_question": "If WebSockets are used for the live monitor, are connections authenticated and are origins strictly validated to prevent CSWSH?",
        "current_state_reply": "Implemented (Aug 2026). Socket.IO CORS restricted to infra.corsOrigins. JWT validated on handshake (auth.token / Bearer). Unauthenticated connections rejected. Broadcast streams scoped to permission rooms (MONITOR_VIEW, ALERTS_VIEW, dashboard for all authenticated users).",
        "plan_of_action": "Implemented. See docs/WFM-Watch-Security-Review-WebSocket-Auth-Response.md.",
        "implementable": "Implemented",
        "owner": "Development",
        "priority": "High",
        "phase": "Phase 2",
        "status": "Implemented",
    },
    {
        "no": 27,
        "security_question": "Are email alerts scrubbed of sensitive raw log data, instead providing a secure link to the dashboard to view details?",
        "current_state_reply": "Notify templates HTML-escape dynamic fields. Escalation emails contain operational data (client codes, counts). Generic alert emails may include unescaped error messages/paths. Emails generally link to dashboard rather than full log dumps.",
        "plan_of_action": "Phase 2: HTML-escape all dynamic email fields. Truncate/sanitize job failure error messages. Ensure no raw log tail content in emails — dashboard link only.",
        "implementable": "Yes (app)",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 2",
        "status": "Planned",
    },
    {
        "no": 28,
        "security_question": "Given Java 8 jjs (Nashorn) is deprecated/removed in modern Java, are there plans to modernize the DB2 connector?",
        "current_state_reply": "Migrated (Aug 2026). DB2 access uses compiled Java 17 JDBC connector (lib/DB2Connector.java → DB2Connector.class, db2jcc4.jar). One JVM per query via java -cp. Nashorn/jjs removed. OpenJDK 17 in prod Docker.",
        "plan_of_action": "Implemented. Optional future: ibm_db native Node driver to eliminate JVM entirely. Full client regression testing before production deploy.",
        "implementable": "Implemented",
        "owner": "Development",
        "priority": "Medium",
        "phase": "Phase 4",
        "status": "Implemented",
    },
]

HEADERS = [
    "#",
    "Security Question / Requirement",
    "Current State / Reply",
    "Plan of Action",
    "Implementable In App?",
    "Owner",
    "Priority",
    "Phase",
    "Status",
]

SUMMARY_ROWS = [
    ("Total review points", "28"),
    ("Implementable in application code", "~20"),
    ("Partially implementable (app + infra/client)", "~4"),
    ("Infrastructure / ops only", "~4"),
    ("", ""),
    ("Phase 1 — Quick wins (days)", "#8 Rate limiting, #9 IDOR permissions, #16 JWT algorithm, #23 Ack audit, #24 Log scrubbing, disable open registration"),
    ("Phase 2 — Medium (1–2 weeks)", "#11 Log path validation, #13 SSH injection fix, #14/#15 JWT revoke + HttpOnly cookies, #26 WebSocket auth, #27 Email scrubbing, #25 Payroll columns"),
    ("Phase 3 — Hardening (2–4 weeks)", "#1 Encrypt all secrets, #2 Key rotation script, #6 TLS config, #19 Polling mins, #20 Timeouts"),
    ("Phase 4 — Strategic", "#28 jjs migration, #7 In-app TOTP MFA (if required)"),
    ("Infra / Ops (not app code)", "#4 Private subnet, #5 Egress firewall, #17 Client sshd restriction, #18 DB2 SELECT grants, #22 SIEM log forwarding"),
]


def style_header(ws, row_num, col_count):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def auto_width(ws, min_width=12, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def main():
    wb = Workbook()

    # --- Main sheet ---
    ws = wb.active
    ws.title = "Security Review"

    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "WFM Watch — Security Review Response & Remediation Plan"
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:I2")
    meta = ws["A2"]
    meta.value = "Generated: July 2026 | Application: WFMControlM (WFM Watch) | ~20 of 28 items implementable in codebase without external API"
    meta.font = Font(italic=True, size=10, color="666666")
    meta.alignment = Alignment(horizontal="center")

    header_row = 4
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(row=header_row, column=col, value=h)
    style_header(ws, header_row, len(HEADERS))

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    for i, row_data in enumerate(ROWS):
        r = header_row + 1 + i
        values = [
            row_data["no"],
            row_data["security_question"],
            row_data["current_state_reply"],
            row_data["plan_of_action"],
            row_data["implementable"],
            row_data["owner"],
            row_data["priority"],
            row_data["phase"],
            row_data["status"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if i % 2 == 1:
                cell.fill = alt_fill

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:I{header_row + len(ROWS)}"

    widths = [5, 45, 45, 45, 18, 22, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Implementation Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")

    for i, (label, value) in enumerate(SUMMARY_ROWS, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws2.cell(row=i, column=2, value=value)
        c.alignment = Alignment(wrap_text=True)

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 90

    # --- Legend sheet ---
    ws3 = wb.create_sheet("Legend")
    legend = [
        ("Implementable In App?", "Yes (app)", "Can be fully implemented in backend/frontend code"),
        ("", "Partial", "App can do part; requires infra, DBA, or client server config"),
        ("", "No — Infra only", "Production/network deployment decision; not application code"),
        ("", "Mostly done", "Already partially implemented; minor enhancements only"),
        ("Priority", "High", "Address in Phase 1–2 before production"),
        ("", "Medium", "Address in Phase 2–3"),
        ("", "Low", "Optional or conditional on other changes"),
        ("Status", "Planned", "Remediation planned, not yet implemented"),
        ("", "Mostly implemented", "Largely in place today"),
        ("", "Ops decision required", "Requires infrastructure/operations team action"),
    ]
    ws3["A1"] = "Field Legend"
    ws3["A1"].font = Font(bold=True, size=12)
    for i, (field, value, desc) in enumerate(legend, start=3):
        ws3.cell(row=i, column=1, value=field)
        ws3.cell(row=i, column=2, value=value)
        ws3.cell(row=i, column=3, value=desc)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 55

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
